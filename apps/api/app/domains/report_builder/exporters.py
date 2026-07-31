from __future__ import annotations

import csv
import io
from typing import Any, Optional

from app.domains.report_builder.base import ReportColumn


def export_csv(columns: list[ReportColumn], rows: list[dict[str, Any]]) -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([c.header for c in columns])
    for row in rows:
        writer.writerow([_format_cell(row.get(c.key, ""), c) for c in columns])
    return output.getvalue()


def export_excel(columns: list[ReportColumn], rows: list[dict[str, Any]], sheet_name: str = "Report") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col.header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, row in enumerate(rows, 2):
        for col_idx, col in enumerate(columns, 1):
            value = _format_cell(row.get(col.key, ""), col)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    for col_idx, col in enumerate(columns, 1):
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = max(
            len(col.header), 12
        )

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_pdf(columns: list[ReportColumn], rows: list[dict[str, Any]], title: str = "Report") -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm, inch
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4) if len(columns) > 6 else A4,
        topMargin=20 * mm,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ReportTitle", parent=styles["Title"], fontSize=16, spaceAfter=12)
    header_style = ParagraphStyle("Header", parent=styles["Normal"], fontSize=8, textColor=colors.white)
    cell_style = ParagraphStyle("Cell", parent=styles["Normal"], fontSize=7)

    elements = [Paragraph(title, title_style), Spacer(1, 6 * mm)]

    header_row = [Paragraph(c.header, header_style) for c in columns]
    data_rows = [[_format_cell(r.get(c.key, ""), c) for c in columns] for r in rows]

    if not data_rows:
        data_rows = [["No data available"] + [""] * (len(columns) - 1)]

    table_data = [header_row] + data_rows

    available_width = (doc.width / mm) * mm
    col_width = available_width / max(len(columns), 1)

    table = RLTable(table_data, colWidths=[col_width] * len(columns), repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )

    elements.append(table)
    doc.build(elements)
    return output.getvalue()


def _format_cell(value: Any, column: ReportColumn) -> Any:
    if value is None:
        return ""
    if column.type == "currency" and isinstance(value, (int, float)):
        return f"{value / 100:,.2f}"
    if column.type == "percentage" and isinstance(value, (int, float)):
        return f"{value:.1f}%"
    if column.type == "date" and hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        return value
    return str(value)
