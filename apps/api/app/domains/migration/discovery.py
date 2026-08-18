from __future__ import annotations

import csv
import io
import json
import logging
import re
from dataclasses import dataclass, field
from typing import Any

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Source readers (D2.2 / Step 2) — CSV, XLSX, JSON (list or keyed), JSONL.
# XLSX is parsed with openpyxl (declared in requirements.txt / pyproject).
# ---------------------------------------------------------------------------


def parse_source(data: bytes, filename: str) -> list[dict[str, Any]]:
    """Parse an uploaded migration file into a list of flat records.

    CSV: first row is the header.  XLSX: first worksheet, first row is the
    header.  JSON: either a list of objects or an object whose first value
    is a list of objects.  JSONL: one object per line.  Returns [] for an
    empty file (never raises for content shape — validation of required
    fields happens later, in the mapping phase).
    """
    name = (filename or "").lower()

    if name.endswith(".xlsx"):
        return _parse_xlsx(data)

    text = _decode(data)
    if name.endswith(".csv"):
        return _parse_csv(text)
    if name.endswith(".jsonl"):
        return _parse_jsonl(text)
    if name.endswith(".json"):
        return _parse_json(text)
    # Unknown extension: attempt CSV, then JSON (best effort).
    try:
        return _parse_csv(text)
    except Exception:
        return _parse_json(text)


def _parse_xlsx(data: bytes) -> list[dict[str, Any]]:
    """Parse an .xlsx workbook into flat records (first sheet only).

    The first non-empty row is the header.  Cells are stringified; the
    transform pipeline (parse_date, normalize_phone, …) normalises values
    later, exactly like the CSV path.
    """
    try:
        import io

        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        logger.warning("Failed to parse XLSX — treating as empty", exc_info=True)
        return []

    try:
        ws = wb.worksheets[0] if wb.worksheets else None
        if ws is None:
            return []

        rows = ws.iter_rows(values_only=True)
        header: list[str] | None = None
        records: list[dict[str, Any]] = []
        for row in rows:
            if row is None:
                continue
            values = ["" if v is None else v for v in row]
            if not any(str(v).strip() for v in values):
                continue
            if header is None:
                header = [str(v).strip() for v in values]
                continue
            record: dict[str, Any] = {}
            for idx, col in enumerate(header):
                if not col:
                    continue
                raw = values[idx] if idx < len(values) else ""
                record[col] = _cell_to_str(raw)
            records.append(record)
        return records
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _cell_to_str(value: Any) -> str:
    """Stringify an openpyxl cell value, keeping dates readable.

    openpyxl yields ``datetime.datetime`` for date cells; a midnight
    datetime normalises to ``YYYY-MM-DD`` (the migration domain stores
    dates, not timestamps), while real timestamps keep their time part.
    """
    import datetime

    if isinstance(value, datetime.datetime):
        if (value.hour, value.minute, value.second, value.microsecond) == (0, 0, 0, 0):
            return value.date().isoformat()
        return value.isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    if value is None:
        return ""
    return str(value).strip()


def _decode(data: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _parse_csv(text: str) -> list[dict[str, Any]]:
    reader = csv.DictReader(io.StringIO(text))
    records: list[dict[str, Any]] = []
    for row in reader:
        cleaned = {str(k).strip(): (v.strip() if isinstance(v, str) else v) for k, v in row.items()}
        records.append(cleaned)
    return records


def _parse_json(text: str) -> list[dict[str, Any]]:
    data = json.loads(text)
    if isinstance(data, list):
        return [r for r in data if isinstance(r, dict)]
    if isinstance(data, dict):
        for value in data.values():
            if isinstance(value, list) and value and isinstance(value[0], dict):
                return [r for r in value if isinstance(r, dict)]
    return []


def _parse_jsonl(text: str) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            parsed = json.loads(line)
            if isinstance(parsed, dict):
                records.append(parsed)
        except json.JSONDecodeError:
            logger.warning("Skipping unparseable JSONL line")
    return records


# ---------------------------------------------------------------------------
# Column profiling (D2.3)
# ---------------------------------------------------------------------------


@dataclass
class ColumnProfile:
    """Deterministic profile of a single source column."""

    name: str
    inferred_type: str
    null_rate: float
    distinct_ratio: float
    is_duplicate_candidate: bool
    sample_values: list[str] = field(default_factory=list)
    looks_like_date: bool = False
    looks_like_email: bool = False
    looks_like_phone: bool = False
    looks_like_identifier: bool = False


_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
_PHONE_RE = re.compile(r"^\+?[\d\s\-().]{7,}$")
_DATE_RE = re.compile(
    r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}$|^\d{1,2}[-/]\d{1,2}[-/]\d{4}$"
    r"|^\d{1,2}[-/]\d{1,2}[-/]\d{2}$"
)
_IDENTIFIER_RE = re.compile(r"(id|no|number|code|ref|roll|adm)", re.IGNORECASE)


def profile_columns(records: list[dict[str, Any]]) -> list[ColumnProfile]:
    """Profile every column present in the source records.

    Null rates and distinct ratios are computed over the full record set
    (cheap, in-memory) so the operator can see at a glance which columns
    are well-populated and which look like duplicate candidates.
    """
    if not records:
        return []

    profiles: list[ColumnProfile] = []
    total = len(records)
    names = list(records[0].keys())

    for name in names:
        values = [r.get(name) for r in records]
        non_null = [v for v in values if v is not None and str(v).strip() != ""]
        null_rate = 1.0 - (len(non_null) / total) if total else 1.0
        distinct = len({str(v) for v in non_null})
        distinct_ratio = (distinct / len(non_null)) if non_null else 0.0

        sample = [str(v)[:60] for v in non_null[:5]]

        looks_like_date = False
        looks_like_email = False
        looks_like_phone = False
        if non_null:
            probe = [str(v) for v in non_null if str(v).strip()]
            looks_like_email = sum(1 for v in probe if _EMAIL_RE.match(v)) >= max(
                1, len(probe) // 2
            )
            looks_like_phone = not looks_like_email and sum(
                1 for v in probe if _PHONE_RE.match(v)
            ) >= max(1, len(probe) // 2)
            looks_like_date = not looks_like_email and sum(
                1 for v in probe if _DATE_RE.match(v)
            ) >= max(1, len(probe) // 2)

        # Type inference over non-null values.
        inferred = _infer_type(non_null)

        profiles.append(
            ColumnProfile(
                name=name,
                inferred_type=inferred,
                null_rate=round(null_rate, 4),
                distinct_ratio=round(distinct_ratio, 4),
                is_duplicate_candidate=distinct_ratio < 0.9 and len(non_null) > 1,
                sample_values=sample,
                looks_like_date=looks_like_date,
                looks_like_email=looks_like_email,
                looks_like_phone=looks_like_phone,
                looks_like_identifier=bool(_IDENTIFIER_RE.search(name)),
            )
        )
    return profiles


def _infer_type(values: list[Any]) -> str:
    if not values:
        return "unknown"
    # A column whose values are all parseable as numbers (and at least one
    # has a decimal point or exponent) is a float; all-integer → int.
    numeric = 0
    floats = 0
    for v in values:
        try:
            float(v)
            numeric += 1
            if "." in str(v) or "e" in str(v).lower():
                floats += 1
        except (TypeError, ValueError):
            pass
    if numeric == len(values):
        return "float" if floats else "int"
    # Pure booleans.
    lowered = {str(v).strip().lower() for v in values}
    if lowered <= {"true", "false", "yes", "no", "1", "0", ""}:
        return "bool"
    return "string"


# ---------------------------------------------------------------------------
# Deterministic mapping suggestions (D2.3 / D2.4)
# ---------------------------------------------------------------------------

#: Canonical SDMAS target fields for the v1 migration (Student).
#: ``source_synonyms`` are the legacy column names we recognise.
#: ``required`` fields block READY when unmapped.
TARGET_FIELDS: dict[str, dict[str, Any]] = {
    "student_number": {
        "label": "Student Number",
        "required": True,
        "synonyms": [
            "student_number",
            "student_no",
            "studentid",
            "student_id",
            "admission_no",
            "adm_no",
            "registration_no",
            "reg_no",
            "roll_no",
            "rollno",
            "enrollment_no",
        ],
    },
    "first_name": {
        "label": "First Name",
        "required": True,
        "synonyms": ["first_name", "firstname", "fname", "given_name", "forename"],
    },
    "last_name": {
        "label": "Last Name",
        "required": True,
        "synonyms": ["last_name", "lastname", "lname", "surname", "family_name"],
    },
    "full_name": {
        "label": "Full Name (split into first/last)",
        "required": False,
        "synonyms": ["full_name", "fullname", "name", "student_name", "studentname"],
        "split_targets": ["first_name", "last_name"],
    },
    "email": {
        "label": "Email",
        "required": False,
        "synonyms": ["email", "email_address", "e_mail", "student_email"],
    },
    "date_of_birth": {
        "label": "Date of Birth",
        "required": False,
        "synonyms": ["dob", "date_of_birth", "birth_date", "birthday", "birthdate"],
    },
    "gender": {
        "label": "Gender",
        "required": False,
        "synonyms": ["gender", "sex"],
        "value_map": {"m": "male", "f": "female", "male": "male", "female": "female"},
    },
    "status": {
        "label": "Status",
        "required": False,
        "synonyms": ["status", "student_status", "enrollment_status"],
    },
    "guardian_phone": {
        "label": "Guardian Phone",
        "required": False,
        "synonyms": ["guardian_phone", "parent_phone", "father_phone", "mother_phone", "phone"],
    },
    # ── Academic structure (Step 2) ────────────────────────────────
    "class_name": {
        "label": "Class / Grade",
        "required": False,
        "synonyms": [
            "class",
            "class_name",
            "grade",
            "grade_level",
            "klass",
            "class_level",
            "form",
            "standard",
            "level",
            "class_grade",
        ],
    },
    "section_name": {
        "label": "Section",
        "required": False,
        "synonyms": ["section", "section_name", "section_no", "sec", "division", "stream"],
    },
    "academic_year_name": {
        "label": "Academic Year",
        "required": False,
        "synonyms": ["academic_year", "academic_year_name", "session", "year", "ay", "batch"],
    },
    # ── Attendance (Step 2) ────────────────────────────────────────
    "attendance_date": {
        "label": "Attendance Date",
        "required": False,
        "synonyms": ["attendance_date", "att_date", "date_of_attendance", "att_date"],
    },
    "attendance_status": {
        "label": "Attendance Status",
        "required": False,
        "synonyms": [
            "attendance_status",
            "status_of_attendance",
            "attendance",
            "att_status",
            "mark",
            "attendance_mark",
        ],
        "value_map": {
            "p": "present",
            "present": "present",
            "a": "absent",
            "absent": "absent",
            "l": "late",
            "late": "late",
            "e": "excused",
            "excused": "excused",
            "h": "holiday",
            "holiday": "holiday",
        },
    },
    # ── Finance (Step 2) ───────────────────────────────────────────
    "amount_paid": {
        "label": "Amount Paid",
        "required": False,
        "synonyms": [
            "amount_paid",
            "fee_paid",
            "paid_amount",
            "amount",
            "fee_amount",
            "paid",
            "payment_amount",
            "fees_paid",
            "fee_collected",
        ],
    },
    "fee_type_name": {
        "label": "Fee Type",
        "required": False,
        "synonyms": ["fee_type", "fee_type_name", "fee_category", "fee_name", "fee_head"],
    },
    "payment_date": {
        "label": "Payment Date",
        "required": False,
        "synonyms": ["payment_date", "paid_on", "date_paid", "receipt_date", "pay_date"],
    },
    "receipt_no": {
        "label": "Receipt Number",
        "required": False,
        "synonyms": [
            "receipt_no",
            "receipt_number",
            "receipt",
            "receipt_id",
            "voucher_no",
            "rcpt_no",
            "rcpt",
        ],
    },
}


@dataclass
class MappingSuggestion:
    """A single inferred source → target mapping."""

    source_field: str
    target_field: str
    confidence: str  # high | medium | low
    reason: str


def suggest_mappings(
    columns: list[str], profiles: list[ColumnProfile] | None = None
) -> list[MappingSuggestion]:
    """Infer deterministic field mappings from source column names.

    Exact synonym match → high confidence.  Normalised fuzzy containment →
    medium.  Nothing matched but the column *looks* like an identifier,
    email, date or phone → low-confidence type-based suggestion.  No AI —
    pure name/type heuristics, fully explainable.
    """
    prof_by_name = {p.name: p for p in (profiles or [])}
    suggestions: list[MappingSuggestion] = []
    matched: set[str] = set()

    def _normalise(name: str) -> str:
        return re.sub(r"[^a-z0-9]", "", name.lower())

    norm_columns = {name: _normalise(name) for name in columns}

    for source in columns:
        norm = norm_columns[source]

        # 1. Exact / fuzzy synonym match.
        # (score, -len(synonym), target, confidence, reason) — score first,
        # then the LONGEST matching synonym wins the tie: a column literally
        # containing "fullname" must map to ``full_name`` (not ``last_name``
        # via the "lname" substring), and "admission_no" to
        # ``student_number`` (not ``receipt_no`` via the bare "no").
        best: tuple[int, int, str, str, str] | None = None
        for target, meta in TARGET_FIELDS.items():
            for synonym in meta["synonyms"]:
                syn_norm = _normalise(synonym)
                if norm == syn_norm:
                    best = (3, 0, target, "high", f"Matches SDMAS field '{meta['label']}'")
                    break
                if syn_norm and syn_norm in norm:
                    # Fuzzy containment (e.g. "parent_phone_number" → phone).
                    score = 2
                    if best is None or (score, -len(syn_norm)) > (best[0], best[1]):
                        best = (
                            score,
                            -len(syn_norm),
                            target,
                            "medium",
                            f"Contains '{synonym}' → {meta['label']}",
                        )
            if best and best[0] == 3:
                break

        if best is None:
            # 2. Type-based fallback.
            prof = prof_by_name.get(source)
            if prof:
                if prof.looks_like_identifier and "student_number" not in norm:
                    best = (
                        1,
                        0,
                        "student_number",
                        "low",
                        "Column looks like an identifier (ID/no/code)",
                    )
                elif prof.looks_like_email:
                    best = (1, 0, "email", "low", "Column contains email-format values")
                elif prof.looks_like_phone:
                    best = (
                        1,
                        0,
                        "guardian_phone",
                        "low",
                        "Column contains phone-format values",
                    )
                elif prof.looks_like_date:
                    best = (1, 0, "date_of_birth", "low", "Column contains date-format values")

        if best is not None:
            suggestions.append(
                MappingSuggestion(
                    source_field=source,
                    target_field=best[2],
                    confidence=best[3],
                    reason=best[4],
                )
            )
            matched.add(source)

    # 3. Unmapped columns — surfaced explicitly so nothing is silently dropped.
    for source in columns:
        if source not in matched:
            suggestions.append(
                MappingSuggestion(
                    source_field=source,
                    target_field="",
                    confidence="low",
                    reason="No SDMAS field detected — map manually or ignore",
                )
            )
    return suggestions


def build_default_mapping(suggestions: list[MappingSuggestion]) -> dict[str, Any]:
    """Turn suggestions into a persisted mapping dict (D2.4).

    Shape: ``{source_field: {target, confidence, reason, transforms: []}}``.
    Only high/medium-confidence suggestions are pre-applied; low-confidence
    and unmapped columns are left for the operator to confirm.
    """
    mapping: dict[str, Any] = {}
    for s in suggestions:
        if not s.target_field or s.confidence == "low":
            continue
        entry = {
            "target": s.target_field,
            "confidence": s.confidence,
            "reason": s.reason,
            "transforms": [],
        }
        meta = TARGET_FIELDS.get(s.target_field)
        if meta and s.target_field == "gender" and meta.get("value_map"):
            entry["transforms"] = [{"op": "map_values", "values": meta["value_map"]}]
        if meta and s.target_field == "attendance_status" and meta.get("value_map"):
            entry["transforms"] = [{"op": "map_values", "values": meta["value_map"]}]
        if meta and s.target_field in ("date_of_birth", "attendance_date", "payment_date"):
            entry["transforms"] = [{"op": "parse_date"}]
        if meta and s.target_field == "guardian_phone":
            entry["transforms"] = [{"op": "normalize_phone"}]
        if meta and s.target_field == "email":
            entry["transforms"] = [{"op": "normalize_email"}]
        mapping[s.source_field] = entry
    return mapping


#: Entity domains and the target fields that signal their presence in a
#: source file (Step 2 — detected entities).  The keys must match the
#: migrator ``entity_type`` names in the engine registry (``fees``, not
#: ``finance``) so the import job can route detected entities to migrators.
ENTITY_TARGETS: dict[str, set[str]] = {
    "students": {
        "student_number",
        "first_name",
        "last_name",
        "full_name",
        "email",
        "date_of_birth",
        "gender",
        "guardian_phone",
    },
    "academic": {"class_name", "section_name", "academic_year_name"},
    "attendance": {"attendance_date", "attendance_status"},
    "fees": {"amount_paid", "fee_type_name", "payment_date", "receipt_no"},
}

#: Human-readable labels for detected entities.
ENTITY_LABELS: dict[str, str] = {
    "students": "Students",
    "academic": "Academic structure",
    "attendance": "Attendance",
    "finance": "Fees / finance",
}


def detect_entities(mapping: dict[str, Any]) -> list[str]:
    """Return the entity domains present in a saved mapping.

    An entity is detected when at least one source column maps to one of
    its target fields.  ``academic`` is stricter: it needs an academic
    year AND a class/section target — a bare "Class" column without a
    year cannot build structure, so it is not an academic import.
    Order is deterministic (dict order of ``ENTITY_TARGETS``).
    """
    targets = {
        str(spec.get("target", "")) for spec in (mapping or {}).values() if isinstance(spec, dict)
    }
    detected: list[str] = []
    for entity, fieldset in ENTITY_TARGETS.items():
        if entity == "academic":
            if "academic_year_name" in targets and targets & {"class_name", "section_name"}:
                detected.append(entity)
        elif targets & fieldset:
            detected.append(entity)
    return detected
